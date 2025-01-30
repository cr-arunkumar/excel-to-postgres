import time
import pandas as pd
import psycopg2
import io
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()


conn_string = f"postgresql://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}@{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
file_path =os.getenv('EXCEL_FILE_PATH') 
file_extension = os.path.splitext(file_path)[1].lower()
CHUNK_SIZE = int(os.getenv('CHUNKS_SIZE'))
NUMBER_OF_ROWS_TO_PROCESS =int(os.getenv('NUMBER_OF_ROWS_TO_PROCESS')) 

def get_total_rows(file_path, file_extension):
    try:
        if file_extension in ['.xlsx', '.xls']:
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            if ws.max_row is None:
                row_count = sum(1 for row in ws.iter_rows())
                return row_count - 1
            return ws.max_row - 1
        else:
            raise ValueError("Unsupported file format")
    except Exception as e:
        print(f"Error getting total rows: {str(e)}")
        return 0

def get_sql_type(dtype):
    if pd.api.types.is_integer_dtype(dtype):
        return 'INTEGER'
    elif pd.api.types.is_float_dtype(dtype):
        return 'DOUBLE PRECISION'
    elif pd.api.types.is_datetime64_any_dtype(dtype):
        return 'TIMESTAMP'
    else:
        return 'TEXT'

def process_chunk(chunk, table_name):
    try:
        with psycopg2.connect(conn_string) as pg_conn:
            with pg_conn.cursor() as cur:
                csv_buffer = io.StringIO()
                chunk.to_csv(csv_buffer, index=False, header=False)
                csv_buffer.seek(0)
                cur.copy_from(csv_buffer, table_name, sep=',', null='')
                pg_conn.commit()
    except Exception as e:
        print(f"Error processing chunk: {str(e)}")
        raise

def main():
    start_time = time.time()
    try:
        table_name = 'copy_test'
        start_time=time.time()
        total_rows =int(NUMBER_OF_ROWS_TO_PROCESS)
        print(f"Total time for getting total rows: {time.time() - start_time:.2f} seconds")
        if file_extension in ['.xlsx', '.xls']:
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows)

            first_chunk = pd.DataFrame(
                [next(rows) for _ in range(CHUNK_SIZE)],
                columns=header
            )
            column_defs = [f'"{col}" {get_sql_type(first_chunk[col].dtype)}' for col in first_chunk.columns]
            column_defs_str = ',\n    '.join(column_defs)

            table_create_sql = f'''
            CREATE TABLE IF NOT EXISTS {table_name} (
                {column_defs_str}
            )
            '''

            with psycopg2.connect(conn_string) as pg_conn:
                with pg_conn.cursor() as cur:
                    cur.execute(table_create_sql)
                    cur.execute(f'TRUNCATE TABLE {table_name}')
                pg_conn.commit()

            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = []
                futures.append(executor.submit(process_chunk, first_chunk, table_name))
                for chunk_start in tqdm(range(0, total_rows, CHUNK_SIZE), desc="Processing chunks"):
                    chunk = pd.DataFrame(
                        [next(rows) for _ in range(CHUNK_SIZE) if _ < total_rows - chunk_start],
                        columns=header
                    )
                    if not chunk.empty:
                        futures.append(executor.submit(process_chunk, chunk, table_name))

                for future in as_completed(futures):
                    future.result()

            print(f"COPY duration: {time.time() - start_time:.2f} seconds")
        else:
            raise ValueError("Unsupported file format. Please use an Excel file.")
    except Exception as e:
        import traceback
        print(f"Error during COPY operation: {str(e)}")
        print("Full traceback:")
        print(traceback.format_exc())

if __name__ == "__main__":
    main()